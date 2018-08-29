import sys, pip, json


def row_to_dict(h, row):
    sub_dict = {}
    for i, cell in enumerate(row):
        sub_dict[h[i]] = cell.value
    return sub_dict



def main():
    wb = openpyxl.load_workbook(sys.argv[1])
    #sheetnames 
    # wb.sheetnames
    sheet = wb["Sheet 1"]
    data = []
    headers = []
    # all_rows(sheet)
    # print(sheet.max_row, sheet.max_column, sheet['B3'].value, sheet['B3'].style)
    # print(sheet['B3'].font)
    # print(sheet['B3'].fill)
    # color cell background
    # sheet['A1'].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")


    # for x in sheet.cell_range.CellRange(max_col = sheet.max_column, max_row = sheet.max_row):
    
    # print(sheet.merged_cells.ranges)
    print(sheet.max_row, sheet.max_column)
    # cells = sheet['A1': 'B6']
    # print(cells)
    # for x,y in cells:
    #     print(x.value, y.value)
    # print(sheet.iter_rows())
    # for row in sheet.iter_rows():
    #     for cell in row:
    #         headers.append(cell.value)
    #     break
    
    for i, row in enumerate(sheet.iter_rows()):
        if i == 0:
            for cell in row:
                headers.append(cell.value)
        else:
            data.append(row_to_dict(headers, row))
        

    # print(headers)
    print(data)

if __name__ == "__main__":
    import openpyxl
    main()
    # try:
    #     import openpyxl
    #     main()
    # except:
    #     print('library not present!')
    

